import sklearn.metrics as sm
import openpyxl as xl
import scipy.io.wavfile
import math
import numpy as np
import ctypes
import os

def sampling(file_audio):
    rate, data = scipy.io.wavfile.read(file_audio)
    data = np.array(data,dtype=np.int16)
    data = np.add(data,[32768])
    return data

def mean_data_sample(data_sample):
    return np.mean(np.power(data_sample,[2]))

def calculate_mse(data_sample, data_stego):
    return sm.mean_squared_error(data_sample,data_stego)

def calculate_snr(data_sample,mse):
    if mse == 0:
        return 'infinite'
    else:
        mds = mean_data_sample(data_sample)
        log_content = mds/mse
        return 10 * math.log(log_content,10)

def calculate_psnr(mse):
    if mse == 0:
        return 'infinite'
    else:
        log_content = (((2 ** 16) - 1) ** 2)/mse
        return 10 * math.log(log_content,10)

def clone_cover_audio(data_sample, filename):
    index_odd = [x for x in range(0, (len(data_sample) * 2) - 1) if x % 2 == 1]
    index_even = [x for x in range(0, (len(data_sample) * 2)) if x % 2 == 0]

    interpolated_sample = np.interp(index_odd, index_even, data_sample)
    interpolated_sample = np.floor(interpolated_sample)

    new_data = []
    i_odd = 0
    i_even = 0
    for x in range (len(data_sample)*2 - 1):
        if x % 2 == 0:
            new_data.append(data_sample[i_even])
            i_even += 1
        else:
            new_data.append(interpolated_sample[i_odd])
            i_odd += 1

    process_data = np.subtract(new_data,[32768])
    process_data = np.array(process_data,dtype=np.int16)
    os.makedirs(os.path.dirname(filename), exist_ok=True)
    scipy.io.wavfile.write(filename, 88200, process_data)
    return new_data

def print_excel(data_mse, data_snr, data_psnr, filename):
    excel = xl.Workbook()
    sheet_mse = excel.create_sheet('Mean Squared Error')

    total_audio = len(data_mse)
    total_payload = len(data_mse[0])

    for index_audio in range (0,total_audio):
        if index_audio == 0:
            for x in range (0, total_payload):
                sheet_mse.cell(row=1,column=x+2).value = 'Payload'+str(x+1)

        for index_payload in range (0,total_payload):
            if index_payload == 0:
                sheet_mse.cell(row=index_audio+2,column=1).value = 'Audio'+str(index_audio+1)
            sheet_mse.cell(row = index_audio+2, column=index_payload+2).value = data_mse[index_audio][index_payload]

    sheet_snr = excel.create_sheet('SNR')

    total_audio = len(data_snr)
    total_payload = len(data_snr[0])

    for index_audio in range(0, total_audio):
        if index_audio == 0:
            for x in range (0, total_payload):
                sheet_snr.cell(row=1,column=x+2).value = 'Payload'+str(x+1)

        for index_payload in range(0, total_payload):
            if index_payload == 0:
                sheet_snr.cell(row=index_audio + 2, column=1).value = 'Audio' + str(index_audio + 1)
            sheet_snr.cell(row=index_audio + 2, column=index_payload + 2).value = data_snr[index_audio][index_payload]

    sheet_psnr = excel.create_sheet('PSNR')

    total_audio = len(data_psnr)
    total_payload = len(data_psnr[0])

    for index_audio in range(0, total_audio):
        if index_audio == 0:
            for x in range (0, total_payload):
                sheet_psnr.cell(row=1,column=x+2).value = 'Payload'+str(x+1)

        for index_payload in range(0, total_payload):
            if index_payload == 0:
                sheet_psnr.cell(row=index_audio + 2, column=1).value = 'Audio' + str(index_audio + 1)
            sheet_psnr.cell(row=index_audio + 2, column=index_payload + 2).value = data_psnr[index_audio][index_payload]

    excel.save(filename)

def main():
    audio = '1'
    payload = '11'

    file_audio = 'dataset/Audio/data'+audio+'_mono.wav'
    filename = 'audio_clone/data_clone_audio'+audio+'.wav'
    file_stego_audio = 'stego_audio/stego_audio'+audio+'_payload'+payload+'/stegoaudio.wav'
    
    sample_audio = sampling(file_audio)
    new_sample_audio = clone_cover_audio(sample_audio,filename)

    sample_stego_audio = sampling(file_stego_audio)

    mse = calculate_mse(new_sample_audio,sample_stego_audio)
    snr = calculate_snr(new_sample_audio,mse)
    psnr = calculate_psnr(mse)

    print("mse  :", mse)
    print("snr  :", snr)
    print("psnr :", psnr)
    ctypes.windll.user32.MessageBoxW(0, "Success","Thesis code", 1)

if __name__ == '__main__':
    main()